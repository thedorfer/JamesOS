from __future__ import annotations

import json
import shutil
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from jamesos.config import VAULT
from jamesos.core.queue import FAILED, IN_PROGRESS, PENDING, PROCESSED, ensure_queue_dirs, enqueue_job

INGEST_ROOT = VAULT / "JamesOS" / "Brain" / "Ingest"
REPORTS = VAULT / "JamesOS" / "Reports" / "Uploads"
TIMELINE_ROOT = VAULT / "JamesOS" / "Timeline"
TEXT_EXTS = {".txt", ".md", ".csv", ".log", ".json", ".xml", ".sql", ".py"}


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _find_manifest(archive_path: str) -> Path | None:
    for manifest_path in INGEST_ROOT.glob("**/derived/manifest.json"):
        try:
            data = _read_json(manifest_path)
        except Exception:
            continue
        if data.get("archive_path") == archive_path:
            return manifest_path
    return None


def _run_command(args: list[str], timeout: int = 45) -> str:
    if shutil.which(args[0]) is None:
        return ""
    try:
        result = subprocess.run(args, capture_output=True, text=True, timeout=timeout)
        return (result.stdout or result.stderr or "").strip()
    except Exception:
        return ""


def _extract_text_from_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)[:20000]
    except Exception:
        pass

    return _run_command(["pdftotext", str(path), "-"], timeout=60)[:20000]


def _extract_docx_text(path: Path) -> str:
    try:
        import docx
        doc = docx.Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)[:20000]
    except Exception:
        pass

    try:
        from xml.etree import ElementTree as ET
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml")
        root = ET.fromstring(xml)
        texts = [node.text for node in root.iter() if node.text]
        return "\n".join(texts)[:20000]
    except Exception:
        return ""


def _extract_spreadsheet(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8", errors="ignore")[:20000]
        lines = text.splitlines()
        return {"text": text, "sheets": [{"name": "csv", "rows_previewed": min(len(lines), 50)}]}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        chunks = []
        for ws in wb.worksheets[:8]:
            rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= 30:
                    break
                rows.append(["" if cell is None else str(cell) for cell in row[:20]])
            sheets.append({"name": ws.title, "rows_previewed": len(rows), "max_row": ws.max_row, "max_column": ws.max_column})
            chunks.append(f"## Sheet: {ws.title}\n" + "\n".join("\t".join(r) for r in rows))
        return {"text": "\n\n".join(chunks)[:20000], "sheets": sheets}
    except Exception as exc:
        return {"text": "", "sheets": [], "error": str(exc)}


def _ocr_image(path: Path) -> str:
    # Uses system tesseract when available. Safe no-op otherwise.
    return _run_command(["tesseract", str(path), "stdout"], timeout=60)[:20000]


def _image_details(path: Path) -> dict[str, Any]:
    details: dict[str, Any] = {}
    try:
        from PIL import Image
        with Image.open(path) as img:
            details.update({"width": img.width, "height": img.height, "format": img.format, "mode": img.mode})
            exif = img.getexif() if hasattr(img, "getexif") else None
            details["has_exif"] = bool(exif)
            if exif:
                details["exif_keys"] = [str(k) for k in list(exif.keys())[:50]]
    except Exception as exc:
        details["error"] = str(exc)
    return details


def _process_by_kind(path: Path, kind: str) -> dict[str, Any]:
    result: dict[str, Any] = {"kind": kind, "processed_at": _now(), "text": "", "summary": "", "metadata": {}, "needs_followup": []}

    if kind == "image":
        result["metadata"]["image"] = _image_details(path)
        result["text"] = _ocr_image(path)
        if result["text"]:
            result["summary"] = "Image processed. OCR text was extracted and saved."
        else:
            result["summary"] = "Image processed. Basic image metadata was saved. OCR/captioning still needs a vision processor."
            result["needs_followup"].extend(["vision caption", "OCR if text is visible"])
        return result

    if kind == "pdf":
        result["text"] = _extract_text_from_pdf(path)
        result["summary"] = "PDF processed. Text was extracted." if result["text"] else "PDF stored. Text extraction needs OCR or another PDF processor."
        if not result["text"]:
            result["needs_followup"].append("OCR scanned PDF")
        return result

    if kind == "document":
        if path.suffix.lower() == ".docx":
            result["text"] = _extract_docx_text(path)
        result["summary"] = "Document processed. Text was extracted." if result["text"] else "Document stored. Text extraction needs an additional parser."
        return result

    if kind == "spreadsheet":
        extracted = _extract_spreadsheet(path)
        result["text"] = extracted.get("text", "")
        result["metadata"]["sheets"] = extracted.get("sheets", [])
        if extracted.get("error"):
            result["metadata"]["error"] = extracted["error"]
        result["summary"] = "Spreadsheet processed. Sheet previews were extracted." if result["text"] else "Spreadsheet stored. Sheet extraction needs an additional parser."
        return result

    if kind == "text" or path.suffix.lower() in TEXT_EXTS:
        result["text"] = path.read_text(encoding="utf-8", errors="ignore")[:20000]
        result["summary"] = "Text file processed and indexed for search."
        return result

    if kind == "audio":
        result["summary"] = "Audio stored. Transcription processor is still needed."
        result["needs_followup"].append("speech-to-text")
        return result

    if kind == "video":
        result["summary"] = "Video stored. Transcript/keyframe processors are still needed."
        result["needs_followup"].extend(["speech-to-text", "keyframe extraction"])
        return result

    result["summary"] = "File stored. Classification-specific processor is still needed."
    result["needs_followup"].append("classification-specific processing")
    return result


def _markdown_result(filename: str, kind: str, result: dict[str, Any], archive_path: str) -> str:
    needs = result.get("needs_followup") or []
    text = result.get("text") or ""
    preview = text[:2000]
    return "\n".join([
        f"# Processed Upload: {filename}",
        "",
        f"Processed: {result.get('processed_at')}",
        f"Kind: {kind}",
        f"Archive: {archive_path}",
        "",
        "## Summary",
        "",
        result.get("summary") or "Processed.",
        "",
        "## Follow-up Needed",
        "",
        *(f"- {item}" for item in needs),
        *( ["- None"] if not needs else [] ),
        "",
        "## Extracted Text Preview",
        "",
        preview or "No text extracted yet.",
        "",
    ])


def _update_manifest(manifest_path: Path, process_result: dict[str, Any], summary_path: Path) -> dict[str, Any]:
    manifest = _read_json(manifest_path)
    manifest["status"] = "processed_initial_pass"
    manifest["processor_result"] = process_result
    manifest["processor_summary_path"] = str(summary_path)
    manifest["updated_at"] = _now()
    _write_json(manifest_path, manifest)
    return manifest


def _write_reports(filename: str, kind: str, result: dict[str, Any], archive_path: str, derived_dir: Path) -> Path:
    summary_path = derived_dir / "processor_summary.md"
    summary_path.write_text(_markdown_result(filename, kind, result, archive_path), encoding="utf-8")
    REPORTS.mkdir(parents=True, exist_ok=True)
    (REPORTS / "Latest Processed Upload.md").write_text(summary_path.read_text(encoding="utf-8"), encoding="utf-8")
    return summary_path


def _append_timeline(filename: str, kind: str, result: dict[str, Any], manifest_path: Path) -> None:
    TIMELINE_ROOT.mkdir(parents=True, exist_ok=True)
    path = TIMELINE_ROOT / f"{datetime.now().strftime('%Y-%m-%d')}.md"
    with path.open("a", encoding="utf-8") as f:
        f.write("\n".join([
            "",
            f"## Processed Upload - {filename}",
            f"- Kind: {kind}",
            f"- Summary: {result.get('summary')}",
            f"- Manifest: {manifest_path}",
            "",
        ]))


def process_attachment_payload(payload: dict[str, Any]) -> dict[str, Any]:
    archive_path = payload.get("path") or ""
    path = Path(archive_path)
    if not path.exists():
        raise FileNotFoundError(archive_path)

    kind = payload.get("kind") or "file"
    filename = payload.get("filename") or path.name
    manifest_path = _find_manifest(str(path))
    if manifest_path is None:
        raise FileNotFoundError(f"No manifest found for {path}")

    derived_dir = manifest_path.parent
    result = _process_by_kind(path, kind)
    result_path = derived_dir / "processor_result.json"
    _write_json(result_path, result)
    summary_path = _write_reports(filename, kind, result, str(path), derived_dir)
    manifest = _update_manifest(manifest_path, result, summary_path)
    _append_timeline(filename, kind, result, manifest_path)

    if result.get("text"):
        enqueue_job("intake", {
            "title": f"Processed Upload - {filename}",
            "content": _markdown_result(filename, kind, result, str(path)),
            "source": "attachment_processor",
            "source_detail": str(summary_path),
        })

    return {"status": "ok", "filename": filename, "kind": kind, "summary": result.get("summary"), "manifest": str(manifest_path), "summary_path": str(summary_path), "manifest_status": manifest.get("status")}


def process_pending_attachment_jobs(limit: int = 10) -> dict[str, Any]:
    ensure_queue_dirs()
    processed = []
    failed = []

    for pending_path in sorted(PENDING.glob("*.json")):
        if len(processed) >= limit:
            break
        job = _read_json(pending_path)
        if job.get("type") != "attachment_process":
            continue

        working_path = IN_PROGRESS / pending_path.name
        shutil.move(str(pending_path), str(working_path))
        try:
            result = process_attachment_payload(job.get("payload") or {})
            job["status"] = "processed"
            job["processed_at"] = _now()
            job["result"] = result
            _write_json(PROCESSED / working_path.name, job)
            working_path.unlink(missing_ok=True)
            processed.append(result)
        except Exception as exc:
            job["status"] = "failed"
            job["failed_at"] = _now()
            job["error"] = str(exc)
            _write_json(FAILED / working_path.name, job)
            working_path.unlink(missing_ok=True)
            failed.append({"job": job.get("id"), "error": str(exc)})

    return {"status": "ok", "processed": len(processed), "failed": len(failed), "results": processed, "errors": failed}
